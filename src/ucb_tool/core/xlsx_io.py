from __future__ import annotations

import hashlib
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.protection import SheetProtection

from ucb_tool import __version__ as TOOL_VERSION
from ucb_tool.core.errors import SchemaError, ValidationError
from ucb_tool.core.ucb_bundle import UcbBundle, UcbInstance

_HEADERS = [
    "Field Path", "Offset (hex)", "Size", "Type", "Raw Bytes",
    "Value", "Decoded", "Enum Options", "Danger", "Help",
]


def _sha256(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def _write_meta(wb: Workbook, bundle: UcbBundle, source_hex: Path) -> None:
    ws = wb.create_sheet("_Meta", 0)
    rows = [
        ("tool_version", TOOL_VERSION),
        ("schema_version", "0.1"),
        ("chip", bundle.chip_id),
        ("family", bundle.family),
        ("source_hex", str(source_hex)),
        ("source_sha256", _sha256(source_hex)),
        ("exported_at", datetime.now(timezone.utc).isoformat()),
    ]
    for i, (k, v) in enumerate(rows, start=1):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 80
    ws.protection = SheetProtection(sheet=True, formatCells=False)


def _write_summary(wb: Workbook, bundle: UcbBundle) -> None:
    ws = wb.create_sheet("Summary", 1)
    headers = ["UCB Name", "Address (ORIG)", "Size", "Confirmation", "CRC OK",
               "Danger Fields"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h).font = Font(bold=True)
    r = 2
    for name, inst in bundle.instances.items():
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=f"0x{inst.orig_addr:08X}")
        ws.cell(row=r, column=3, value=inst.schema.size)
        ws.cell(row=r, column=4, value="?")
        ws.cell(row=r, column=5, value="?")
        danger_fields = [
            f.path for f in inst.fields
            if f.schema.get("x-danger", "safe") != "safe"
        ]
        ws.cell(row=r, column=6, value=", ".join(danger_fields))
        r += 1
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.protection = SheetProtection(sheet=True, formatCells=False)


def _write_ucb_sheet(wb: Workbook, inst: UcbInstance) -> None:
    ws = wb.create_sheet(inst.schema.name)
    for c, h in enumerate(_HEADERS, start=1):
        ws.cell(row=1, column=c, value=h).font = Font(bold=True)
    r = 2
    for f in inst.fields:
        if f.schema.get("x-bitfield"):
            continue  # skip parent display row; child bits listed separately
        raw = bytes(inst.buf_orig[f.offset:f.offset + f.size])
        try:
            val: int | None = inst.get(f.path)
        except Exception:  # noqa: BLE001
            val = None
        ws.cell(row=r, column=1, value=f.path)
        ws.cell(row=r, column=2, value=f"0x{f.offset:04X}")
        ws.cell(row=r, column=3, value=f.size)
        ws.cell(row=r, column=4, value=f.schema.get("type", ""))
        ws.cell(row=r, column=5, value=raw.hex())
        render = f.schema.get("x-render")
        if val is not None and render == "hex":
            ws.cell(row=r, column=6, value=f"0x{val:X}")
        else:
            ws.cell(row=r, column=6, value=val)
        names = f.schema.get("x-enum-names") or {}
        if val is not None and names:
            ws.cell(row=r, column=7, value=names.get(str(val), ""))
        ws.cell(row=r, column=8, value=", ".join(f"{k}={v}" for k, v in names.items()))
        ws.cell(row=r, column=9, value=f.danger)
        ws.cell(row=r, column=10, value=f.schema.get("x-help", ""))
        r += 1
    for col in range(1, len(_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.protection = SheetProtection(sheet=True, formatCells=False)
    # Unlock the Value column so users can edit it in Excel.
    for row in ws.iter_rows(min_row=2, min_col=6, max_col=6):
        for cell in row:
            cell.protection = Protection(locked=False)


def export_to_xlsx(bundle: UcbBundle, out_path: str | Path, source_hex: Path) -> None:
    wb = Workbook()
    # Remove the default 'Sheet' that openpyxl creates.
    default_name = wb.active.title if wb.active else None
    if default_name and default_name in wb.sheetnames:
        del wb[default_name]
    _write_meta(wb, bundle, source_hex)
    _write_summary(wb, bundle)
    for inst in bundle.instances.values():
        _write_ucb_sheet(wb, inst)
    wb.save(str(out_path))


def _parse_value(raw: object, schema: dict[str, Any]) -> int | None:
    if raw is None or raw == "":
        return None
    if isinstance(raw, bool):
        # bool is a subclass of int in Python; treat as int.
        return int(raw)
    if isinstance(raw, (int, float)):
        return int(raw)
    if isinstance(raw, str):
        s = raw.strip()
        if s == "":
            return None
        names = schema.get("x-enum-names") or {}
        rev = {v: int(k) for k, v in names.items()}
        if s in rev:
            return rev[s]
        try:
            return int(s, 0)
        except ValueError:
            pass
    raise ValidationError("<value>", f"cannot parse {raw!r}")


def apply_xlsx(
    bundle: UcbBundle,
    xlsx_path: str | Path,
    *,
    lenient: bool = False,
) -> None:
    """Read an edited .xlsx back into `bundle` in place.

    Strict mode (default) rejects unknown sheets and unknown field-path rows.
    Lenient mode ignores them but still enforces schema_version match.
    """
    wb = load_workbook(str(xlsx_path), data_only=True)
    sheets = set(wb.sheetnames)

    if "_Meta" not in sheets:
        raise SchemaError(f"{xlsx_path}: missing _Meta sheet")
    if "Summary" not in sheets:
        raise SchemaError(f"{xlsx_path}: missing Summary sheet")

    known_ucbs = set(bundle.instances.keys())
    expected = known_ucbs | {"_Meta", "Summary"}
    unknown = sheets - expected
    if unknown and not lenient:
        raise SchemaError(f"{xlsx_path}: unknown sheets {sorted(unknown)}")

    meta_ws = wb["_Meta"]
    meta: dict[object, object] = {}
    for r in meta_ws.iter_rows(min_row=1, max_col=2):
        meta[r[0].value] = r[1].value
    if meta.get("schema_version") != "0.1":
        raise SchemaError(
            f"schema_version mismatch: expected 0.1, got "
            f"{meta.get('schema_version')!r}"
        )

    for name in known_ucbs:
        if name not in sheets:
            continue
        ws = wb[name]
        inst = bundle.instances[name]
        by_path = {f.path: f for f in inst.fields}
        for r in range(2, ws.max_row + 1):
            path_cell = ws.cell(row=r, column=1).value
            value_cell = ws.cell(row=r, column=6).value
            if path_cell is None:
                continue
            if path_cell not in by_path:
                if lenient:
                    continue
                raise SchemaError(
                    f"{xlsx_path} sheet {name}: unknown field path "
                    f"{path_cell!r}"
                )
            f = by_path[path_cell]
            if f.read_only or f.computed:
                continue
            parsed = _parse_value(value_cell, f.schema)
            if parsed is not None:
                inst.set(path_cell, parsed)


def diff_bundles(a: UcbBundle, b: UcbBundle, out_path: str | Path) -> int:
    """Write a 'Changes' sheet to out_path. Returns count of differing fields."""
    wb = Workbook()
    default_name = wb.active.title if wb.active else None
    if default_name and default_name in wb.sheetnames:
        del wb[default_name]
    ws = wb.create_sheet("Changes")
    headers = ["UCB.Field", "A value", "B value", "Danger"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h).font = Font(bold=True)
    r = 2
    count = 0
    for name in set(a.instances) | set(b.instances):
        ia = a.instances.get(name)
        ib = b.instances.get(name)
        if ia is None or ib is None:
            ws.cell(row=r, column=1, value=f"{name} (only in one side)")
            r += 1
            count += 1
            continue
        for f in ia.fields:
            if f.read_only:
                continue
            try:
                va = ia.get(f.path)
                vb = ib.get(f.path)
            except Exception:  # noqa: BLE001
                continue
            if va != vb:
                ws.cell(row=r, column=1, value=f"{name}.{f.path}")
                ws.cell(row=r, column=2, value=va)
                ws.cell(row=r, column=3, value=vb)
                ws.cell(row=r, column=4, value=f.danger)
                r += 1
                count += 1
    wb.save(str(out_path))
    return count
