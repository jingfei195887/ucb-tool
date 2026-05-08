from pathlib import Path

import pytest
from openpyxl import load_workbook

import ucb_tool
from ucb_tool.core.hex_io import write_hex
from ucb_tool.gui.main_window import MainWindow

SCHEMAS = Path(ucb_tool.__file__).parent / "schemas"


@pytest.fixture
def prepared_win(qtbot, qapp, tmp_path):
    hex_path = tmp_path / "u.hex"
    data = {0xAF400000 + i: 0xFF for i in range(256)}
    write_hex(hex_path, data)

    win = MainWindow()
    qtbot.addWidget(win)

    # Programmatically load — bypass File/Open dialog
    win._source_path = hex_path
    win._current_chip = "tc4d9"
    win._bundle = win._load(str(hex_path), "tc4d9")
    win._populate_tree()
    win.action_save.setEnabled(True)
    return win, hex_path, tmp_path


def test_apply_xlsx_action(prepared_win, monkeypatch):
    win, hex_path, tmp_path = prepared_win
    from ucb_tool.core.xlsx_io import export_to_xlsx
    xlsx_path = tmp_path / "u.xlsx"
    export_to_xlsx(win._bundle, xlsx_path, source_hex=hex_path)

    wb = load_workbook(xlsx_path)
    ws = wb["BMHD_0"]
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "STAD":
            ws.cell(row=r, column=6, value="0x80000000")
            break
    wb.save(xlsx_path)

    # Patch file dialog + danger dialog
    monkeypatch.setattr(
        "ucb_tool.gui.main_window.QFileDialog.getOpenFileName",
        staticmethod(lambda *a, **kw: (str(xlsx_path), "")),
    )
    monkeypatch.setattr(
        "ucb_tool.gui.dialogs.danger_confirm.DangerConfirmDialog.exec",
        lambda self: int(self.DialogCode.Accepted),
    )
    win.on_apply_xlsx()
    assert win._bundle["BMHD_0"].get("STAD") == 0x80000000
