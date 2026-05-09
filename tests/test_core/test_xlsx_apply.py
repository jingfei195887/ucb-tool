from pathlib import Path

import pytest
from openpyxl import load_workbook

from tests.conftest import LEGACY_COMMON_DIR
from ucb_tool.core.errors import SchemaError
from ucb_tool.core.hex_io import write_hex
from ucb_tool.core.ucb_bundle import UcbBundle
from ucb_tool.core.xlsx_io import apply_xlsx, export_to_xlsx


def _make_bundle(tmp_path: Path) -> tuple[Path, UcbBundle]:
    hex_path = tmp_path / "u.hex"
    data = {0xAF400000 + i: 0xFF for i in range(256)}
    write_hex(hex_path, data)
    bundle = UcbBundle.load(hex_path, "tc4d9",
                            common_dirs=[LEGACY_COMMON_DIR],
                            chip_schema_dir=None)
    return hex_path, bundle


def test_apply_edits_value_column_back(tmp_path):
    hex_path, bundle = _make_bundle(tmp_path)
    xlsx_path = tmp_path / "u.xlsx"
    export_to_xlsx(bundle, xlsx_path, source_hex=hex_path)

    wb = load_workbook(xlsx_path)
    ws = wb["BMHD_0"]
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "STAD":
            ws.cell(row=r, column=6, value="0x80000000")
            break
    wb.save(xlsx_path)

    apply_xlsx(bundle, xlsx_path)
    assert bundle["BMHD_0"].get("STAD") == 0x80000000


def test_apply_rejects_unknown_sheet(tmp_path):
    hex_path, bundle = _make_bundle(tmp_path)
    xlsx_path = tmp_path / "u.xlsx"
    export_to_xlsx(bundle, xlsx_path, source_hex=hex_path)

    wb = load_workbook(xlsx_path)
    wb.create_sheet("IMPOSTER")
    wb.save(xlsx_path)

    with pytest.raises(SchemaError):
        apply_xlsx(bundle, xlsx_path)


def test_apply_rejects_unknown_field_row(tmp_path):
    hex_path, bundle = _make_bundle(tmp_path)
    xlsx_path = tmp_path / "u.xlsx"
    export_to_xlsx(bundle, xlsx_path, source_hex=hex_path)

    wb = load_workbook(xlsx_path)
    ws = wb["BMHD_0"]
    new_row = ws.max_row + 1
    ws.cell(row=new_row, column=1, value="GHOST_FIELD")
    ws.cell(row=new_row, column=6, value=1)
    wb.save(xlsx_path)

    with pytest.raises(SchemaError):
        apply_xlsx(bundle, xlsx_path)


def test_apply_accepts_enum_label_or_int(tmp_path):
    hex_path, bundle = _make_bundle(tmp_path)
    xlsx_path = tmp_path / "u.xlsx"
    export_to_xlsx(bundle, xlsx_path, source_hex=hex_path)

    wb = load_workbook(xlsx_path)
    ws = wb["BMHD_0"]
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "BMI.HWCFG":
            ws.cell(row=r, column=6, value="ASC bootstrap")
            break
    wb.save(xlsx_path)

    apply_xlsx(bundle, xlsx_path)
    assert bundle["BMHD_0"].get("BMI.HWCFG") == 3


def test_apply_lenient_ignores_unknown(tmp_path):
    hex_path, bundle = _make_bundle(tmp_path)
    xlsx_path = tmp_path / "u.xlsx"
    export_to_xlsx(bundle, xlsx_path, source_hex=hex_path)

    wb = load_workbook(xlsx_path)
    wb.create_sheet("EXTRA")
    wb.save(xlsx_path)

    apply_xlsx(bundle, xlsx_path, lenient=True)  # should not raise
