from openpyxl import load_workbook

from tests.conftest import LEGACY_COMMON_DIR
from ucb_tool.core.hex_io import write_hex
from ucb_tool.core.ucb_bundle import UcbBundle
from ucb_tool.core.xlsx_io import export_to_xlsx


def test_export_creates_meta_summary_and_per_ucb_sheets(tmp_path):
    hex_path = tmp_path / "u.hex"
    data = {0xAF400000 + i: 0xFF for i in range(256)}
    write_hex(hex_path, data)
    bundle = UcbBundle.load(hex_path, "tc4d9",
                            common_dirs=[LEGACY_COMMON_DIR],
                            chip_schema_dir=None)

    out = tmp_path / "u.xlsx"
    export_to_xlsx(bundle, out, source_hex=hex_path)
    wb = load_workbook(out)
    assert "_Meta" in wb.sheetnames
    assert "Summary" in wb.sheetnames
    assert "BMHD_0" in wb.sheetnames

    meta = wb["_Meta"]
    meta_dict = {row[0].value: row[1].value for row in meta.iter_rows(min_row=1, max_col=2)}
    assert meta_dict["chip"] == "tc4d9"
    assert meta_dict["source_hex"].endswith("u.hex")
    assert len(meta_dict["source_sha256"]) == 64

    bmhd = wb["BMHD_0"]
    assert bmhd.cell(row=1, column=1).value == "Field Path"
    col1 = [bmhd.cell(row=r, column=1).value for r in range(2, bmhd.max_row + 1)]
    assert "STAD" in col1
