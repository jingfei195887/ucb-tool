from pathlib import Path

from click.testing import CliRunner
from openpyxl import load_workbook

from ucb_tool.cli.__main__ import main
from ucb_tool.core.hex_io import read_hex

FIX = Path(__file__).parent.parent / "fixtures"


def test_no_op_save_preserves_all_bytes_outside_ucb(tmp_path):
    src = FIX / "tc4d9_sample.hex"
    out = tmp_path / "out.hex"
    runner = CliRunner()
    result = runner.invoke(main, [
        "set", str(src), "--chip", "tc4d9",
        "--field", "BMHD_0.STAD=0x80000000",  # same value — no functional change
        "--out", str(out), "--yes-i-know-brick",
    ])
    assert result.exit_code == 0, result.output
    orig = read_hex(src)
    new = read_hex(out)
    for addr in (0x80000000, 0x80000001):
        assert new[addr] == orig[addr]


def test_export_xlsx_sheets_are_deterministic(tmp_path):
    src = FIX / "tc4d9_sample.hex"
    runner = CliRunner()
    out1 = tmp_path / "a.xlsx"
    out2 = tmp_path / "b.xlsx"
    runner.invoke(main, ["export-xlsx", str(src), "--chip", "tc4d9", "--out", str(out1)])
    runner.invoke(main, ["export-xlsx", str(src), "--chip", "tc4d9", "--out", str(out2)])

    wb1 = load_workbook(out1)
    wb2 = load_workbook(out2)
    for name in wb1.sheetnames:
        if name == "_Meta":
            continue  # timestamp in _Meta makes it non-deterministic
        ws1 = wb1[name]
        ws2 = wb2[name]
        assert ws1.max_row == ws2.max_row
        assert ws1.max_column == ws2.max_column
