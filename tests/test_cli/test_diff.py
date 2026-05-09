from click.testing import CliRunner
from openpyxl import load_workbook

from tests.conftest import LEGACY_COMMON_DIR
from ucb_tool.cli.__main__ import main
from ucb_tool.core.hex_io import write_hex


def test_diff_produces_change_rows(tmp_path):
    a = tmp_path / "a.hex"
    b = tmp_path / "b.hex"
    data_a = {0xAF400000 + i: 0xFF for i in range(256)}
    data_b = dict(data_a)
    data_b[0xAF400000] = 0x00  # change low byte of STAD
    write_hex(a, data_a)
    write_hex(b, data_b)

    runner = CliRunner()
    out = tmp_path / "diff.xlsx"
    result = runner.invoke(main, [
        "diff", str(a), str(b), "--chip", "tc4d9", "--out", str(out),
        "--schemas", str(LEGACY_COMMON_DIR),
    ])
    assert result.exit_code == 0, result.output

    wb = load_workbook(out)
    assert "Changes" in wb.sheetnames
    ws = wb["Changes"]
    col1 = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    assert any(v and "STAD" in str(v) for v in col1)
