from pathlib import Path

from click.testing import CliRunner
from openpyxl import load_workbook

from tests.conftest import LEGACY_COMMON_DIR
from ucb_tool.cli.__main__ import main
from ucb_tool.core.hex_io import read_hex, slice_range, write_hex

_SCH = ["--schemas", str(LEGACY_COMMON_DIR)]


def test_apply_xlsx_round_trip_with_consent(tmp_path: Path):
    src = tmp_path / "u.hex"
    data = {0xAF400000 + i: 0xFF for i in range(256)}
    write_hex(src, data)

    runner = CliRunner()
    xlsx_path = tmp_path / "u.xlsx"
    result = runner.invoke(main, [
        "export-xlsx", str(src), "--chip", "tc4d9",
        "--out", str(xlsx_path), *_SCH,
    ])
    assert result.exit_code == 0

    wb = load_workbook(xlsx_path)
    ws = wb["BMHD_0"]
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "STAD":
            ws.cell(row=r, column=6, value="0x80000000")
            break
    wb.save(xlsx_path)

    out_hex = tmp_path / "new.hex"
    # brick danger → need --yes-i-know-brick
    result = runner.invoke(main, [
        "apply-xlsx", str(src),
        "--chip", "tc4d9",
        "--xlsx", str(xlsx_path),
        "--out", str(out_hex),
        "--yes-i-know-brick",
        *_SCH,
    ])
    assert result.exit_code == 0, result.output
    assert slice_range(read_hex(out_hex), 0xAF400000, 4) == b"\x00\x00\x00\x80"
